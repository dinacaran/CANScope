from __future__ import annotations

import array
import csv

import numpy as np
import pytest

from core.export import ExportService, ExportTimebase
from core.signal_store import SignalSeries


def _series(name: str, timestamps, values) -> SignalSeries:
    return SignalSeries(
        channel=1,
        message_name=f"{name}Message",
        message_id=1,
        signal_name=name,
        unit="",
        timestamps=array.array("d", timestamps),
        values=array.array("d", values),
    )


def _read_csv(path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as stream:
        return list(csv.reader(stream))


def test_csv_reference_signal_uses_its_exact_timestamps(tmp_path):
    reference = _series("Reference", [0.05, 0.15, 0.35], [5.0, 15.0, 35.0])
    other = _series("Other", [0.0, 0.1, 0.2, 0.4], [0.0, 10.0, 20.0, 40.0])
    timebase = ExportTimebase.from_reference_signal(reference.key)
    output = tmp_path / "reference.csv"

    ExportService.export_series_to_csv(
        [reference, other], output, timebase=timebase
    )

    rows = _read_csv(output)
    assert [float(row[0]) for row in rows[1:]] == [0.05, 0.15, 0.35]
    assert [float(row[2]) for row in rows[1:]] == [0.0, 10.0, 20.0]
    assert ExportService.count_data_rows(
        [reference, other], timebase=timebase
    ) == 3


def test_csv_manual_recurrence_builds_regular_timestamp_axis(tmp_path):
    first = _series("First", [0.0, 0.2, 0.4], [0.0, 2.0, 4.0])
    second = _series("Second", [0.1, 0.3], [10.0, 30.0])
    timebase = ExportTimebase.from_recurrence(0.1)
    output = tmp_path / "regular.csv"

    ExportService.export_series_to_csv(
        [first, second], output, timebase=timebase
    )

    rows = _read_csv(output)
    np.testing.assert_allclose(
        [float(row[0]) for row in rows[1:]],
        [0.0, 0.1, 0.2, 0.3, 0.4],
    )
    assert [row[2] for row in rows[1:]] == ["", "10.0", "10.0", "30.0", "30.0"]


def test_excel_uses_selected_manual_recurrence(tmp_path):
    from openpyxl import load_workbook

    signal = _series("Signal", [0.0, 0.25], [1.0, 2.0])
    output = tmp_path / "regular.xlsx"

    ExportService.export_series_to_excel(
        [signal],
        output,
        timebase=ExportTimebase.from_recurrence(0.1),
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[0] == ("Time", "Signal")
    np.testing.assert_allclose([row[0] for row in rows[1:]], [0.0, 0.1, 0.2])
    assert [row[1] for row in rows[1:]] == [1.0, 1.0, 1.0]


def test_large_manual_recurrence_count_and_excel_cap_do_not_build_full_grid(tmp_path):
    from openpyxl import load_workbook

    signal = _series("Signal", [0.0, 100.0], [1.0, 2.0])
    timebase = ExportTimebase.from_recurrence(5e-6)
    assert ExportService.count_data_rows([signal], timebase=timebase) == 20_000_001

    output = tmp_path / "capped.xlsx"
    ExportService.export_series_to_excel(
        [signal],
        output,
        timebase=timebase,
        max_data_rows=3,
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert len(rows) == 4
    np.testing.assert_allclose([row[0] for row in rows[1:]], [0.0, 5e-6, 10e-6])


def test_typical_reference_recurrence_uses_median_positive_interval():
    signal = _series("Signal", [0.0, 0.01, 0.02, 0.12], [1.0, 2.0, 3.0, 4.0])

    assert ExportService.typical_recurrence_seconds(signal) == pytest.approx(0.01)


def test_invalid_or_missing_timebase_is_rejected():
    signal = _series("Signal", [0.0, 1.0], [1.0, 2.0])

    with pytest.raises(ValueError, match="greater than zero"):
        ExportService.count_data_rows(
            [signal], timebase=ExportTimebase.from_recurrence(0.0)
        )
    with pytest.raises(ValueError, match="not plotted"):
        ExportService.count_data_rows(
            [signal],
            timebase=ExportTimebase.from_reference_signal("CH1::Missing::Signal"),
        )
