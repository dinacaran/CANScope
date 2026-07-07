"""
Diagnostic-doc lookup — deliberately simple, no vector DB.

Markdown files live under ``config/diagnostics/knowledge/<platform>/*.md``.  A
file may begin with a small header of ``key: value`` lines (``dtc:`` /
``signals:``) followed by free prose.  Retrieval is pure-Python TF-IDF cosine
over the documents, with a keyword-overlap fallback — no numpy/DB dependency.

Excel diagnostic manuals (``*.xlsx``) are also first-class knowledge sources —
see the "xlsx ingestion" section below. They are discovered from both
``knowledge/<platform>/`` and the ``knowledge/`` root (shared across all
platforms), parsed row-by-row into the same ``KnowledgeDoc`` corpus, and
never need manual conversion to Markdown.

Input:  a DTC code or fault name / finding title.
Output: the most relevant doc snippets plus candidate signal names to probe.
"""
from __future__ import annotations

import logging
import math
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from core.diagnostics.config_loader import default_config_dir

_log = logging.getLogger(__name__)

_TOKEN_RE = re.compile(r"[A-Za-z0-9_]+")
_BACKTICK_RE = re.compile(r"`([A-Za-z0-9_]+)`")


def _tokenize(text: str) -> list[str]:
    return [t.lower() for t in _TOKEN_RE.findall(text)]


@dataclass(slots=True)
class KnowledgeDoc:
    name: str
    text: str
    dtcs: list[str] = field(default_factory=list)
    signals: list[str] = field(default_factory=list)
    _tf: Counter = field(default_factory=Counter)


@dataclass(slots=True)
class DocSnippet:
    name: str
    text: str
    score: float
    signals: list[str] = field(default_factory=list)
    dtcs: list[str] = field(default_factory=list)


def default_knowledge_dir(platform: str) -> Path:
    return default_config_dir() / "knowledge" / platform


def _parse_doc(path: Path) -> KnowledgeDoc:
    raw = path.read_text(encoding="utf-8")
    lines = raw.splitlines()
    dtcs: list[str] = []
    signals: list[str] = []

    # Leading ``key: value`` header block (stops at first blank/non-header line).
    body_start = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            body_start = i + 1
            break
        m = re.match(r"^([A-Za-z_]+)\s*:\s*(.+)$", stripped)
        if not m:
            body_start = i
            break
        key, val = m.group(1).lower(), m.group(2)
        items = [v.strip() for v in re.split(r"[,;]", val) if v.strip()]
        if key == "dtc":
            dtcs += items
        elif key in ("signal", "signals"):
            signals += items
        body_start = i + 1

    body = "\n".join(lines[body_start:]).strip()
    # Signals also gleaned from inline `Backticked` identifiers.
    signals += _BACKTICK_RE.findall(raw)
    signals = list(dict.fromkeys(signals))  # dedupe, keep order

    doc = KnowledgeDoc(name=path.stem, text=body or raw, dtcs=dtcs, signals=signals)
    doc._tf = Counter(_tokenize(raw))
    return doc


# ── xlsx ingestion ───────────────────────────────────────────────────────
#
# Excel diagnostic manuals are parsed generically by header keywords, never
# by sheet name, so a differently-arranged manual still yields usable docs.
# Three recognized shapes:
#   * dtc_manual        — a DTC/Measurement_ID column + explanatory columns
#                          (root cause, troubleshooting, ...) -> one doc/row.
#   * can_diagnosis      — a Rule ID column + a signal-condition column
#                          -> one doc/row, condition kept verbatim.
#   * signal_dictionary — a Signal Name column (no DTC/Rule ID)
#                          -> one combined doc for the whole sheet.
# Anything else (presentation sheets, raw per-sample data dumps) is skipped.

_DTC_CODE_RE = re.compile(r"[A-Za-z]\d{3,5}")
_DTC_NUMERIC_RE = re.compile(r"\b\d{2,6}\b")

_EXPLANATORY_KEYWORDS = (
    "fault title", "detection", "monitor logic", "root cause", "troubleshoot",
    "repair", "validation", "fault reaction", "consistency",
    "symptom", "immediate check",
)
# NOTE: deliberately no bare "limp" keyword — "Fault Reaction / Limp Strategy"
# is already caught by "fault reaction", and a bare "limp" substring-matches
# raw signal columns like "LimpHomeMode" in time-series data dumps, which
# would wrongly classify them as explanatory content.

#: path (resolved, str) -> (mtime, parsed docs). Reparsed only if mtime changes.
_XLSX_CACHE: dict[str, tuple[float, list[KnowledgeDoc]]] = {}


def _normalize_header(header) -> str:
    text = "" if header is None else str(header)
    text = re.sub(r"[_/\-]", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def _find_col(norm_headers: list[str], predicate) -> int | None:
    for i, n in enumerate(norm_headers):
        if predicate(n):
            return i
    return None


def _classify_sheet(norm_headers: list[str]) -> str:
    has_dtc = any("dtc" in n for n in norm_headers)
    has_measurement_id = any("measurement id" in n for n in norm_headers)
    has_rule_id = any("rule id" in n for n in norm_headers)
    has_signal_name = any(
        "signal" in n and "name" in n and "plot" not in n and "condition" not in n
        for n in norm_headers
    )
    has_condition = any("signal" in n and "condition" in n for n in norm_headers)
    has_timestamp = any(n == "timestamp" for n in norm_headers)
    has_explanatory = any(
        any(kw in n for kw in _EXPLANATORY_KEYWORDS) for n in norm_headers
    )

    if has_timestamp and not has_explanatory:
        return "skip"
    if has_signal_name and not has_dtc and not has_rule_id:
        return "signal_dictionary"
    if has_rule_id and (has_condition or has_dtc or has_measurement_id):
        return "can_diagnosis"
    if (has_dtc or has_measurement_id) and has_explanatory:
        return "dtc_manual"
    return "skip"


def _extract_dtc_forms(*values) -> list[str]:
    """Pull both DTC forms (e.g. ``P0522`` and ``522``) out of cell values."""
    out: list[str] = []
    for v in values:
        if v is None:
            continue
        text = str(v)
        out += _DTC_CODE_RE.findall(text)
        out += _DTC_NUMERIC_RE.findall(text)
    return list(dict.fromkeys(out))  # dedupe, keep order


def _split_signals(value) -> list[str]:
    if value is None:
        return []
    return [s.strip() for s in re.split(r"[,;]", str(value)) if s.strip()]


def _row_is_empty(row: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _row_body_lines(headers: list[str], row: list) -> list[str]:
    """Every non-empty cell becomes a ``Header: value`` line, column order.

    This is what lets unknown/unlisted columns show up automatically instead
    of being dropped, and what carries labeled fields (Fault Title, Detection
    / Monitor Logic, Likely Root Causes, ...) and rule conditions verbatim
    into the doc text without a fixed per-field mapping.
    """
    lines: list[str] = []
    for header, value in zip(headers, row):
        if not header or value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        lines.append(f"{header}: {text}")
    return lines


def _read_sheet_rows(ws) -> tuple[list[str], list[list]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    data_rows = [list(r) for r in rows_iter]
    return headers, data_rows


def _docs_from_dtc_manual_sheet(
    sheet_name: str, workbook_stem: str,
    headers: list[str], rows: list[list], norm: list[str],
) -> list[KnowledgeDoc]:
    dtc_idx = _find_col(norm, lambda n: "dtc" in n)
    mid_idx = _find_col(norm, lambda n: "measurement id" in n)
    sig_idx = _find_col(norm, lambda n: "signal" in n and "plot" in n)

    docs: list[KnowledgeDoc] = []
    for row in rows:
        if _row_is_empty(row):
            continue
        dtc_val = row[dtc_idx] if dtc_idx is not None else None
        mid_val = row[mid_idx] if mid_idx is not None else None
        dtcs = _extract_dtc_forms(dtc_val, mid_val)
        if not dtcs:
            continue
        signals = _split_signals(row[sig_idx]) if sig_idx is not None else []
        lines = _row_body_lines(headers, row)
        if not lines:
            continue
        key = dtc_val or mid_val or dtcs[0]
        doc = KnowledgeDoc(
            name=f"{workbook_stem}::{sheet_name}::{key}",
            text="\n".join(lines), dtcs=dtcs, signals=signals,
        )
        doc._tf = Counter(_tokenize(doc.text))
        docs.append(doc)
    return docs


def _docs_from_can_diagnosis_sheet(
    sheet_name: str, workbook_stem: str,
    headers: list[str], rows: list[list], norm: list[str],
) -> list[KnowledgeDoc]:
    rule_idx = _find_col(norm, lambda n: "rule id" in n)
    dtc_idx = _find_col(norm, lambda n: "dtc" in n or "measurement id" in n)
    sig_idx = _find_col(norm, lambda n: "signal" in n and "plot" in n)

    docs: list[KnowledgeDoc] = []
    for row in rows:
        if _row_is_empty(row):
            continue
        rule_id = row[rule_idx] if rule_idx is not None else None
        dtc_val = row[dtc_idx] if dtc_idx is not None else None
        dtcs = _extract_dtc_forms(dtc_val)
        signals = _split_signals(row[sig_idx]) if sig_idx is not None else []
        lines = _row_body_lines(headers, row)
        if not lines:
            continue
        key = rule_id or (dtcs[0] if dtcs else None)
        if key is None:
            continue
        doc = KnowledgeDoc(
            name=f"{workbook_stem}::{sheet_name}::{key}",
            text="\n".join(lines), dtcs=dtcs, signals=signals,
        )
        doc._tf = Counter(_tokenize(doc.text))
        docs.append(doc)
    return docs


def _doc_from_signal_dictionary_sheet(
    sheet_name: str, workbook_stem: str,
    headers: list[str], rows: list[list], norm: list[str],
) -> KnowledgeDoc | None:
    name_idx = _find_col(norm, lambda n: "signal" in n and "name" in n)

    blocks: list[str] = []
    signals: list[str] = []
    for row in rows:
        if _row_is_empty(row):
            continue
        lines = _row_body_lines(headers, row)
        if not lines:
            continue
        blocks.append("\n".join(lines))
        if name_idx is not None and row[name_idx]:
            sig = str(row[name_idx]).strip()
            if sig and sig not in signals:
                signals.append(sig)

    if not blocks:
        return None
    text = "\n\n".join(blocks)
    doc = KnowledgeDoc(
        name=f"{workbook_stem}::{sheet_name}", text=text, dtcs=[], signals=signals,
    )
    doc._tf = Counter(_tokenize(text))
    return doc


def _parse_xlsx(path: Path) -> list[KnowledgeDoc]:
    try:
        import openpyxl
    except ImportError as exc:
        _log.warning("Knowledge: openpyxl not available, skipping %s (%s)", path.name, exc)
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        _log.warning("Knowledge: could not open %s (%s) — skipping.", path.name, exc)
        return []

    docs: list[KnowledgeDoc] = []
    skipped: list[str] = []
    try:
        stem = path.stem
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                headers, rows = _read_sheet_rows(ws)
            except Exception as exc:
                skipped.append(f"{sheet_name} (unreadable: {exc})")
                continue
            if not headers:
                skipped.append(f"{sheet_name} (empty)")
                continue

            norm = [_normalize_header(h) for h in headers]
            kind = _classify_sheet(norm)

            if kind == "skip":
                skipped.append(
                    f"{sheet_name} (no DTC/Rule ID/Signal Name column with "
                    "usable content — presentation sheet or raw data dump)"
                )
                continue
            if kind == "signal_dictionary":
                doc = _doc_from_signal_dictionary_sheet(sheet_name, stem, headers, rows, norm)
                if doc is not None:
                    docs.append(doc)
                else:
                    skipped.append(f"{sheet_name} (signal dictionary shape, no usable rows)")
            elif kind == "can_diagnosis":
                new_docs = _docs_from_can_diagnosis_sheet(sheet_name, stem, headers, rows, norm)
                if new_docs:
                    docs.extend(new_docs)
                else:
                    skipped.append(f"{sheet_name} (rule shape, no usable rows)")
            elif kind == "dtc_manual":
                new_docs = _docs_from_dtc_manual_sheet(sheet_name, stem, headers, rows, norm)
                if new_docs:
                    docs.extend(new_docs)
                else:
                    skipped.append(f"{sheet_name} (DTC shape, no usable rows)")
    except Exception as exc:
        _log.warning("Knowledge: error parsing %s (%s) — using what was parsed so far.", path.name, exc)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not docs:
        _log.warning(
            "Knowledge: %s yielded no docs. Skipped sheets: %s",
            path.name, "; ".join(skipped) or "(none)",
        )
    return docs


def _docs_from_xlsx_cached(path: Path) -> list[KnowledgeDoc]:
    """Parse *path* once per session; reparse only if its mtime changes."""
    try:
        mtime = path.stat().st_mtime
    except OSError as exc:
        _log.warning("Knowledge: cannot stat %s (%s) — skipping.", path, exc)
        return []

    key = str(path.resolve())
    cached = _XLSX_CACHE.get(key)
    if cached is not None and cached[0] == mtime:
        return cached[1]

    docs = _parse_xlsx(path)
    _XLSX_CACHE[key] = (mtime, docs)
    return docs


class KnowledgeIndex:
    """TF-IDF index over one platform's knowledge docs."""

    def __init__(self, docs: list[KnowledgeDoc]) -> None:
        self.docs = docs
        self._idf: dict[str, float] = {}
        self._build_idf()

    # ── construction ────────────────────────────────────────────────────
    @classmethod
    def build(cls, platform: str, knowledge_dir: Path | None = None) -> "KnowledgeIndex":
        base = knowledge_dir or default_knowledge_dir(platform)
        docs: list[KnowledgeDoc] = []
        if base.is_dir():
            for path in sorted(base.glob("*.md")):
                try:
                    docs.append(_parse_doc(path))
                except Exception:
                    continue

        # .xlsx manuals: platform folder + the shared knowledge/ root.
        xlsx_dirs = [base]
        if base.parent.is_dir() and base.parent != base:
            xlsx_dirs.append(base.parent)
        for d in xlsx_dirs:
            if not d.is_dir():
                continue
            for path in sorted(d.glob("*.xlsx")):
                if path.name.startswith("~$"):
                    continue
                docs.extend(_docs_from_xlsx_cached(path))

        return cls(docs)

    def _build_idf(self) -> None:
        n = len(self.docs)
        if n == 0:
            return
        df: Counter = Counter()
        for doc in self.docs:
            for term in set(doc._tf):
                df[term] += 1
        for term, d in df.items():
            self._idf[term] = math.log((1 + n) / (1 + d)) + 1.0

    # ── retrieval ───────────────────────────────────────────────────────
    def retrieve(self, query: str, k: int = 3) -> list[DocSnippet]:
        """Return up to *k* doc snippets most relevant to *query*."""
        if not self.docs:
            return []
        q_tf = Counter(_tokenize(query))
        # Boost exact DTC / signal matches so a DTC code reliably retrieves its doc.
        q_terms = set(q_tf)
        scored: list[tuple[float, KnowledgeDoc]] = []
        for doc in self.docs:
            score = self._cosine(q_tf, doc)
            for dtc in doc.dtcs:
                if dtc.lower() in query.lower():
                    score += 5.0
            if q_terms & {s.lower() for s in doc.signals}:
                score += 1.0
            if score > 0:
                scored.append((score, doc))
        scored.sort(key=lambda t: t[0], reverse=True)
        out: list[DocSnippet] = []
        for score, doc in scored[:k]:
            out.append(DocSnippet(
                name=doc.name,
                text=_snippet(doc.text),
                score=round(score, 4),
                signals=list(doc.signals),
                dtcs=list(doc.dtcs),
            ))
        return out

    def candidate_signals(self, query: str, k: int = 3) -> list[str]:
        """Flatten candidate signal names from the top-k docs for *query*."""
        seen: list[str] = []
        for snip in self.retrieve(query, k):
            for s in snip.signals:
                if s not in seen:
                    seen.append(s)
        return seen

    def _cosine(self, q_tf: Counter, doc: KnowledgeDoc) -> float:
        if not q_tf or not doc._tf:
            return 0.0
        dot = 0.0
        for term, qc in q_tf.items():
            idf = self._idf.get(term)
            if idf is None:
                continue
            dot += (qc * idf) * (doc._tf.get(term, 0) * idf)
        q_norm = math.sqrt(sum((c * self._idf.get(t, 0.0)) ** 2 for t, c in q_tf.items()))
        d_norm = math.sqrt(sum((c * self._idf.get(t, 0.0)) ** 2 for t, c in doc._tf.items()))
        if q_norm == 0 or d_norm == 0:
            return 0.0
        return dot / (q_norm * d_norm)


def _snippet(text: str, max_chars: int = 800) -> str:
    text = text.strip()
    return text if len(text) <= max_chars else text[:max_chars].rsplit(" ", 1)[0] + " …"
